#!/usr/bin/env python3
"""Build an "actuals" hours report from a Basecamp time-tracking export.

Reads a raw Basecamp time log (the `pdftotext -layout` output of the export, or
any text with one "DATE  PERSON  HOURS  DESCRIPTION" entry per line), buckets
every hour by ADO ticket (row) and by each person's role (Dev/QA/PM column),
and writes an actuals-only report.

The role of each person is NOT inferred — you must pass it in via --roles. The
caller is responsible for asking who is Dev/QA/PM; this script refuses to guess.

Aggregation uses only the Python standard library and always writes a CSV. The
.xlsx is written too when `openpyxl` is importable (`pip install openpyxl`).

Usage:
  build_actuals.py --log timelog.txt \\
      --roles '{"Elena Dotsenko":"QA","Matt McClain":"PM","Sean Dillon":"Dev"}' \\
      --estimation separate \\
      --out actuals.xlsx

  --estimation {separate,exclude,general}  how to treat SOW/estimation lines
      separate (default): their own "SOW Review & Estimation" row
      exclude:            dropped from the report entirely
      general:            folded into the General Hours row
  --features '{"1509":"Auto Membership Transfer", ...}'  optional ADO->name map
  --general-label "General Hours"   label for the general/meetings row
"""
import argparse
import csv
import json
import os
import re
import sys
from collections import defaultdict

ROLES = ("Dev", "QA", "PM")
ENTRY = re.compile(
    r"^\s*([A-Z][a-z]{2}\s+\d{1,2})\s+"      # date, e.g. "Jul 08"
    r"([A-Za-z.'-]+(?:\s+[A-Za-z.'-]+)+?)\s+"  # person (>=2 words)
    r"(\d+(?:\.\d+)?)\s+"                       # hours
    r"(.+?)\s*$"                                # description
)
TICKET = re.compile(r"\b(\d{3,4})\b")
# non-ticket lines whose text marks them as future-SOW scoping/estimation work
ESTIMATION_KW = ("estimat", "eval and")


def parse_log(path):
    """Yield (date, person, hours, desc) for every recognizable entry line."""
    rows = []
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            m = ENTRY.match(line.rstrip("\n"))
            if not m:
                continue
            date, person, hours, desc = m.groups()
            # skip the "TOTAL   116.00" footer (no person / not a real entry)
            if person.strip().upper() == "TOTAL":
                continue
            rows.append((date, person.strip(), float(hours), desc.strip()))
    return rows


def classify(desc):
    """Return ('ticket', '1509') | ('estimation', None) | ('general', None)."""
    m = TICKET.search(desc)
    if m:
        return "ticket", m.group(1)
    low = desc.lower()
    if any(k in low for k in ESTIMATION_KW):
        return "estimation", None
    return "general", None


def build(rows, roles, estimation_mode):
    """Aggregate into {row_key: {role: hours}} plus diagnostics."""
    unknown = sorted({p for _, p, _, _ in rows if p not in roles})
    if unknown:
        raise SystemExit(
            "Role not provided for: "
            + ", ".join(unknown)
            + "\nEvery person must be mapped to Dev/QA/PM via --roles "
            "(ask the caller — do not guess)."
        )

    tickets = defaultdict(lambda: defaultdict(float))
    general = defaultdict(float)
    estimation = defaultdict(float)
    dropped = 0.0

    for _, person, hours, desc in rows:
        role = roles[person]
        kind, ticket = classify(desc)
        if kind == "ticket":
            tickets[ticket][role] += hours
        elif kind == "estimation":
            if estimation_mode == "exclude":
                dropped += hours
            elif estimation_mode == "general":
                general[role] += hours
            else:  # separate
                estimation[role] += hours
        else:
            general[role] += hours

    return tickets, general, estimation, dropped


def make_table(tickets, general, estimation, features, general_label):
    """Build the ordered list of output rows (dicts)."""
    header = ["ADO", "Feature", "Actual Dev", "Actual QA", "Actual PM",
              "Actual Total"]
    out = []

    def line(ado, feature, buckets):
        dev, qa, pm = buckets.get("Dev", 0.0), buckets.get("QA", 0.0), buckets.get("PM", 0.0)
        return {"ADO": ado, "Feature": feature, "Actual Dev": dev,
                "Actual QA": qa, "Actual PM": pm, "Actual Total": dev + qa + pm}

    for ado in sorted(tickets, key=int):
        out.append(line(ado, features.get(ado, ""), tickets[ado]))
    if estimation:
        out.append(line("", "SOW Review & Estimation", estimation))
    if general:
        out.append(line("", general_label, general))

    totals = {"ADO": "", "Feature": "TOTAL"}
    for col in ("Actual Dev", "Actual QA", "Actual PM", "Actual Total"):
        totals[col] = round(sum(r[col] for r in out), 2)
    out.append(totals)
    return header, out


def write_csv(path, header, rows):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=header)
        w.writeheader()
        for r in rows:
            w.writerow({k: (round(v, 2) if isinstance(v, float) else v)
                        for k, v in r.items()})


def write_xlsx(path, header, rows):
    """Write the .xlsx. Excel owns the arithmetic: each row's Actual Total is a
    =SUM(Dev:PM) formula, and the TOTAL row sums each column's cells. The script
    never bakes a total into a cell value."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False
    # Column layout: A ADO, B Feature, C Dev, D QA, E PM, F Total
    DEV, QA, PM, TOT = "C", "D", "E", "F"
    wb = Workbook()
    ws = wb.active
    ws.title = "Actuals"
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="D9D9D9")

    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    data_rows = [r for r in rows if r.get("Feature") != "TOTAL"]
    total_row = next((r for r in rows if r.get("Feature") == "TOTAL"), None)

    first = 2  # first data row in the sheet
    for r in data_rows:
        rn = ws.max_row + 1
        ws.append([r["ADO"], r["Feature"], r["Actual Dev"], r["Actual QA"],
                   r["Actual PM"], f"=SUM({DEV}{rn}:{PM}{rn})"])
        for c in ws[rn]:
            c.border = border
        for c in ws[rn][2:]:
            c.number_format = "0.00"
    last = ws.max_row  # last data row

    if total_row is not None and data_rows:
        rn = ws.max_row + 1
        ws.append(["", "TOTAL"] + [f"=SUM({col}{first}:{col}{last})"
                                   for col in (DEV, QA, PM, TOT)])
        for c in ws[rn]:
            c.border = border
            c.font = Font(bold=True)
        for c in ws[rn][2:]:
            c.number_format = "0.00"

    widths = [8, 42, 12, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    return True


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--log", required=True, help="raw time-log text file")
    ap.add_argument("--roles", required=True,
                    help='JSON {"Full Name":"Dev|QA|PM"} for EVERY person')
    ap.add_argument("--estimation", choices=["separate", "exclude", "general"],
                    default="separate")
    ap.add_argument("--features", default="{}",
                    help='optional JSON {"1509":"Feature name"}')
    ap.add_argument("--general-label", default="General Hours")
    ap.add_argument("--out", default="actuals.xlsx", help="output .xlsx path")
    args = ap.parse_args()

    roles = json.loads(args.roles)
    bad = {p: r for p, r in roles.items() if r not in ROLES}
    if bad:
        raise SystemExit(f"Roles must be one of {ROLES}; got {bad}")
    features = json.loads(args.features)

    rows = parse_log(args.log)
    if not rows:
        raise SystemExit(f"No time entries parsed from {args.log}")
    tickets, general, estimation, dropped = build(rows, roles, args.estimation)
    header, table = make_table(tickets, general, estimation, features,
                               args.general_label)

    csv_path = os.path.splitext(args.out)[0] + ".csv"
    write_csv(csv_path, header, table)
    xlsx_ok = write_xlsx(args.out, header, table)

    logged = round(sum(h for _, _, h, _ in rows), 2)
    reported = table[-1]["Actual Total"]
    print(f"Parsed {len(rows)} entries, {logged} h logged.")
    print(f"Reported {reported} h" + (f" (+{dropped} h excluded estimation)"
          if dropped else ""))
    if round(reported + dropped, 2) != logged:
        print(f"WARNING: reconciliation mismatch "
              f"({reported} + {dropped} != {logged})", file=sys.stderr)
    print(f"CSV : {csv_path}")
    print(f"XLSX: {args.out}" if xlsx_ok
          else "XLSX: skipped (pip install openpyxl to enable)")


if __name__ == "__main__":
    main()
